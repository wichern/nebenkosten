#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import copy
from dataclasses import dataclass
import datetime
from typing import List, Tuple

import openpyxl

__author__ = "Paul Wichern"
__license__ = "MIT"
__version__ = "0.0.1"

# TODO: Add table for meter values
# TODO: Stimmt "share amount" bei Consumption based?
# TODO: Add units to table view
# TODO: BillItem contains strings with placeholder for 'row' and then ResultSheet.write() will replace that with the actual row
# TODO: Print report, showing;
#           - missing files for invoices
#           - missing meter values
#           - missing bills for a category
# TODO: In "Menge" column, append the unit (kWh, Tage, m³)
# TODO: Handle conversion of m³ into kWh
# TODO: Janine: beautify example bill
# TODO: Bundle Excel Sheet with all invoices into a zip
# TODO: Create PDF of the first Excel page with a method that will do nothing if pypdf2 is not available.
# TODO: Guess NK for 2023
# TODO: Make front page
# TODO: Mention invoice filename in BillItem Row
# TODO: Mention why bill item was split in "Anmerkung"

def warning(msg: str):
    print('WARNUNG: ' + msg)

class MeterValueException(Exception):
    pass

@dataclass
class Date:
    ''' A date (Converts from string representation(s)) '''
    date: datetime.date

    def __init__(self, date: datetime.date):
        self.date = date

    @classmethod
    def from_str(cls, date_str: str):
        return Date(datetime.datetime.strptime(date_str, '%d.%m.%Y').date())

    def yesterday(self):
        return Date(self.date - datetime.timedelta(days=1))

    def tomorrow(self):
        return Date(self.date + datetime.timedelta(days=1))

    def __le__(self, other) -> bool:
        return self.date <= other.date

    def __lt__(self, other) -> bool:
        return self.date < other.date

    def __gt__(self, other) -> bool:
        return self.date > other.date

    def __str__(self):
        return self.date.strftime('%d.%m.%Y')

@dataclass
class DateRange:
    ''' A date range '''
    begin: Date
    end: Date

    def overlaps(self, other) -> bool:
        ''' Check if this date overlaps with given date '''
        return self.begin in other or self.end in other

    def __contains__(self, date: Date) -> bool:
        ''' Check if date is in range '''
        return date >= self.begin and date <= self.end

@dataclass
class Invoice:
    ''' Invoice structure '''
    type: str
    supplier: str
    invoice_number: str
    date: Date
    notes: str
    range: DateRange
    net: str
    amount: str
    tax: str

    def from_row(row):
        ''' Convert table row into Invoice structure '''
        return Invoice(
            row[0],
            row[1],
            row[2],
            Date.from_str(row[3]),
            row[4],
            DateRange(Date.from_str(row[5]), Date.from_str(row[6])),
            row[7],
            row[8],
            row[9])

@dataclass
class Appartement:
    ''' Appartement structure '''
    name: str
    size: str

    def from_row(row):
        ''' Convert table row into Appartement structure '''
        return Appartement(row[0], row[1])

@dataclass
class Tenant:
    ''' Tenant structure '''
    name: str
    appartement: str
    moving_in: datetime
    moving_out: datetime
    people: int

    @classmethod
    def from_row(cls, row):
        ''' Convert table row into Tenant structure '''
        # For 'moving out' we also accept None
        moving_out = None
        if row[3]:
            moving_out = Date.from_str(row[3])
        return Tenant(row[0], row[1], Date.from_str(row[2]), moving_out, int(row[4]))

    def __contains__(self, date: Date) -> bool:
        ''' Check if date is in range '''
        if date < self.moving_in:
            return False
        if self.moving_out:
            return date <= self.moving_out
        return True

@dataclass
class MeterValue:
    ''' MeterValue structure '''
    name: str
    count: str
    date: Date
    notes: str

    def from_row(row):
        ''' Convert table row into MeterValue structure '''
        return MeterValue(row[0], row[1], Date.from_str(row[2]), row[3])

@dataclass
class BillCalculationItem:
    ''' BillCalculationItem structure '''
    appartement: str
    split: str
    unit: str
    invoice_type: str
    meter: str

    def from_row(row):
        ''' Convert table row into BillCalculationItem structure '''
        return BillCalculationItem(row[0], row[1], row[2], row[3], row[4])

@dataclass
class BillItem:
    ''' A row in the bill '''
    range: DateRange
    days: str
    type: str
    description: str
    meter: str
    net: str
    amount: str
    tax: str
    gross: str
    invoice_range_days: str
    share_name: str
    share_percentage: str
    unit: str
    sum: str

    def __init__(self, invoice: Invoice, bci: BillCalculationItem, appartements: List[Appartement], bill_range: DateRange, row_idx: int, total_people_count: int, appartement_size: str, consumption: str, tenant: Tenant):
        self.range = DateRange(
            max(bill_range.begin, invoice.range.begin), 
            min(bill_range.end, invoice.range.end))
        self.days = '=_xlfn.days(B{0}, A{0})'.format(row_idx)
        self.type = invoice.type
        self.description = invoice.notes
        self.meter = bci.meter
        self.net = invoice.net
        self.amount = invoice.amount
        self.tax = invoice.tax
        self.gross = '=G{0}*H{0}*(1+I{0})'.format(row_idx)
        self.invoice_range_days = f'=_xlfn.days("{invoice.range.end}", "{invoice.range.begin}")'
        self.share_name = bci.split
        self.share_percentage = self.__get_share_percentage(appartements, bci.split, total_people_count, appartement_size, consumption, tenant)
        self.unit = str(bci.unit)

        if bci.split == 'Nach Verbrauch':
            self.sum = '=G{0}*(1+I{0})*M{0}'.format(row_idx)
        else:
            self.sum = '=J{0}/K{0}*C{0}*M{0}'.format(row_idx)

    def __get_share_percentage(self, appartements: List['Appartement'], split: str, total_people_count: int, appartement_size: str, consumption: str, tenant: Tenant):
        if split == 'Pro Wohnung':
            return f'=1/{len(appartements)}'
        elif split == 'Pro Person':
            return f'={tenant.people}/{total_people_count}'
        elif split == 'Pro Quadratmeter':
            return f'={appartement_size}/{sum([a.size for a in appartements])}'
        elif split == 'Nach Verbrauch':
            return '=' + consumption
        elif split == 'Hälfte':
            return '=1/2'
        elif split == 'Drittel':
            return '=1/3'
        elif split == 'Viertel':
            return '=1/4'
        elif split == 'Komplett':
            return '1'
        else:
            raise InvalidCellValue('Unknown bill split: "%s"' % split)

class ResultSheet(object):
    '''
    '''

    template_filepath = 'example-bill.xlsx'

    def __init__(self, filepath):
        self._filepath = filepath

    def write(self, row: int, bill_item: BillItem):
        for c in range(1, 15):
            self._sheet.cell(row=row, column=c).style = 'content'
        self._sheet.cell(row=row, column=1).value = bill_item.range.begin.date
        self._sheet.cell(row=row, column=1).number_format = 'DD.MM.YY'
        self._sheet.cell(row=row, column=2).value = bill_item.range.end.date
        self._sheet.cell(row=row, column=2).number_format = 'DD.MM.YY'
        self._sheet.cell(row=row, column=3).value = bill_item.days
        self._sheet.cell(row=row, column=4).value = bill_item.type
        self._sheet.cell(row=row, column=5).value = bill_item.description
        self._sheet.cell(row=row, column=6).value = bill_item.meter
        self._sheet.cell(row=row, column=7).value = bill_item.net
        self._sheet.cell(row=row, column=7).number_format = '0.00" "€'
        self._sheet.cell(row=row, column=8).value = bill_item.amount
        self._sheet.cell(row=row, column=9).value = bill_item.tax
        self._sheet.cell(row=row, column=9).number_format = '0.00" "%'
        self._sheet.cell(row=row, column=10).value = bill_item.gross
        self._sheet.cell(row=row, column=10).number_format = '0.00" "€'
        self._sheet.cell(row=row, column=11).value = bill_item.invoice_range_days
        self._sheet.cell(row=row, column=11).number_format = '0" Tage"'
        self._sheet.cell(row=row, column=12).value = bill_item.share_name

        self._sheet.cell(row=row, column=13).value = bill_item.share_percentage
        if bill_item.share_name == 'Pro Person':
            self._sheet.cell(row=row, column=13).number_format = '0.00" "%'
        elif bill_item.share_name == 'Nach Verbrauch':
            self._sheet.cell(row=row, column=13).number_format = '0.00" ' + bill_item.unit + '"'
        elif bill_item.share_name == 'Pro Wohnung':
            self._sheet.cell(row=row, column=13).number_format = '0.00" "%'
        elif bill_item.share_name == 'Pro Quadratmeter':
            self._sheet.cell(row=row, column=13).number_format = '0.00" "%'

        self._sheet.cell(row=row, column=14).value = bill_item.sum
        self._sheet.cell(row=row, column=14).number_format = '0.00" "€'

    def __enter__(self):
        ''' Open the template sheet '''
        self._wb = openpyxl.load_workbook(self.template_filepath)
        self._sheet = self._wb['Details']

        header = openpyxl.styles.NamedStyle(name='header')
        header.font = openpyxl.styles.Font(name='Calibri', bold=True, size=11)
        self._wb.add_named_style(header)

        content = openpyxl.styles.NamedStyle(name='content')
        content.font = openpyxl.styles.Font(name='Calibri', size=11)
        self._wb.add_named_style(content)

        # Apply styles
        for cell in self._sheet['1']:
            cell.style = 'header'

        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        ''' Write the result '''

        self._wb.save(self._filepath)

def read_table(path):
    ''' Read workbook '''

    workbook = openpyxl.load_workbook(filename=str(path))
    invoices = []
    appartements = []
    tenants = []
    meter_values = []
    bcis = []

    sheet = workbook['Rechnungen']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            invoices.append(Invoice.from_row(row))
        else:
            break

    # for i in invoices:
    #     print(i)

    sheet = workbook['Wohnungen']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            appartements.append(Appartement.from_row(row))
        else:
            break

    sheet = workbook['Mieter']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            tenants.append(Tenant.from_row(row))
        else:
            break

    sheet = workbook['Zählerstände']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            meter_values.append(MeterValue.from_row(row))
        else:
            break

    sheet = workbook['Abrechnungseinstellungen']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            bcis.append(BillCalculationItem.from_row(row))
        else:
            break

    return invoices, appartements, tenants, meter_values, bcis

def get_people_count_change_dates(tenants: List[Tenant], range: DateRange) -> List[Tuple[Date, int]]:
    ret: List['Date', int] = []
    date: 'Date' = range.begin
    idx: int = 0
    people_count_before: int = -1
    while date <= range.end:
        people_count: int = sum([t.people for t in tenants if date in t])

        if people_count != people_count_before:
            ret.append((date, people_count))
        
        people_count_before = people_count
        date = date.tomorrow()
        idx += 1
    return ret

def split_invoice_where_person_count_changes(split_dates: List[Tuple[Date, int]], invoice: Invoice, range: DateRange) -> List[Tuple[Invoice, int]]:
    ''' Split all invoices at dates, where the total people count over all tenants changed. '''

    ret: List[Tuple['Invoice', int]] = []

    people_count_before: int = 0
    invoice_begin: 'Date' = invoice.range.begin
    end_of_invoice: bool = False
    for split_date, people_count in split_dates:
        if split_date == range.begin:
            people_count_before = people_count
            continue

        if split_date >= invoice.range.begin:
            invoice_split = copy.deepcopy(invoice)
            invoice_split.range.begin = max(invoice.range.begin, invoice_begin)
            invoice_split.range.end = min(invoice.range.end, split_date.yesterday())
            ret.append((invoice_split, people_count_before))

            invoice_begin = split_date

        people_count_before = people_count

        if split_date > invoice.range.end:
            end_of_invoice = True
            break

    if not end_of_invoice:
        invoice_split = copy.deepcopy(invoice)
        invoice_split.range.begin = invoice_begin
        ret.append((invoice_split, people_count_before))
    
    return ret

def get_meter_value(meter_values: List[MeterValue], meter_name: str, date: Date) -> MeterValue:
    mv_start: MeterValue = None
    mv_end: MeterValue = None
    for meter_value in meter_values:
        if meter_value.date == date:
            return meter_value
        if meter_value.date < date:
            mv_start = meter_value
        else:
            mv_end = meter_value
            break
    
    if not mv_start:
        raise MeterValueException('Kein Zählerstand für "{0}" vor dem {1} gefunden.'.format(meter_name, str(date)))
    if not mv_end:
        raise MeterValueException('Kein Zählerstand für "{0}" nach dem {1} gefunden.'.format(meter_name, str(date)))

    warning('Keinen Zählerstand für "{0}" am "{1}" gefunden. Stand wird geschätzt.'.format(meter_name, str(date)))

    return MeterValue(
        meter_name,
        f'({mv_end.count}-{mv_start.count})/_xlfn.days("{str(mv_end.date)}", "{str(mv_start.date)}")*_xlfn.days("{str(date)}", "{str(mv_start.date)}")',
        date,
        'Geschätzt'
    )

def get_consumption(meter_name: str, meter_values: List[MeterValue], range: DateRange) -> str:
    if not meter_name:
        raise MeterValueException('Kein Zähler für Abrechungseinstellung definiert, obwohl nach Verbrauch abgerechnet werden soll.')

    # Filter meter values by name.
    meter_values = list(filter(lambda mv: mv.name == meter_name, meter_values))

    # Sort meter values by date.
    meter_values = sorted(meter_values, key=lambda mv: mv.date)

    # Get starting meter value
    mv_start = get_meter_value(meter_values, meter_name, range.begin)
    mv_end = get_meter_value(meter_values, meter_name, range.end)

    return f'{mv_end.count}-{mv_start.count}'
        
def main(args):
    invoices, appartements, tenants, meter_values, bcis = read_table(args.invoices)

    bill_range = DateRange(args.begin, args.end)
    split_dates = get_people_count_change_dates(tenants, bill_range)

    appartement_size: str = next(a for a in appartements if a.name == args.appartement).size
    tenant = next(t for t in tenants if bill_range.begin in t)

    print('Mieter: ' + tenant.name)

    # Splitting invoices
    # For each BCI of the appartement
    with ResultSheet('test.xlsx') as result_sheet:
        row: int = 2
        for bci in filter(lambda bci: bci.appartement == args.appartement, bcis):
            # Get all invoices in the date range matching our BCI type.
            for invoice in filter(lambda i: i.type == bci.invoice_type and i.range.overlaps(bill_range), invoices):
                if bci.split == 'Pro Person':
                    for invoice_split, people_count in split_invoice_where_person_count_changes(split_dates, invoice, invoice.range):
                        bill_item = BillItem(invoice_split, bci, appartements, bill_range, row, people_count, appartement_size, 0, tenant)
                        result_sheet.write(row, bill_item)
                        row += 1
                elif bci.split == 'Nach Verbrauch':
                    consumption_range: DateRange = DateRange(
                        max(invoice.range.begin, bill_range.begin),
                        min(invoice.range.end, bill_range.end))
                    consumption: str = get_consumption(bci.meter, meter_values, consumption_range)
                    bill_item = BillItem(invoice, bci, appartements, bill_range, row, 0, appartement_size, consumption, None)
                    result_sheet.write(row, bill_item)
                    row += 1
                else:
                    bill_item = BillItem(invoice, bci, appartements, bill_range, row, 0, appartement_size, 0, None)
                    result_sheet.write(row, bill_item)
                    row += 1


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Nebenkosten Abrechner')

    parser.add_argument('invoices', help='Pfad zu Nebenkostentabelle', nargs='?')
    parser.add_argument('begin', help='Startdatum', nargs='?', 
        type=lambda s: Date.from_str(s))
    parser.add_argument('end', help='Enddatum', nargs='?',
        type=lambda s: Date.from_str(s))
    parser.add_argument('appartement', help='Wohnung', nargs='?')

    args = parser.parse_args()
    main(args)