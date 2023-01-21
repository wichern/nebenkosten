#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
Reader for input and output excel sheets.
'''

import enum
import logging
from typing import List

import openpyxl

from nebenkosten.types import Invoice, Appartement, Tenant, MeterValue, Meter
from nebenkosten.types import Date, DateRange, BillCalculationItem
from nebenkosten import InputFileError, SplitType

class InputSheet(enum.Enum):
    ''' Name of the sheets in the input file '''
    INVOICES = 'Rechnungen'
    APPARTEMENTS = 'Wohnungen'
    TENANTS = 'Mieter'
    METER_VALUES = 'Zählerstände'
    METERS = 'Zähler'
    BILL_CALCULATION_ITEMS = 'Abrechnungseinstellungen'

# pylint: disable=too-many-instance-attributes
class InputSheetReader:
    '''
    Parse the input sheet.
    '''

    def __init__(self, path, appartement_name: str, bill_range: DateRange):
        '''
        :param path:                Path to the input file
        :param appartement_name:    Name of the appartement which we are creating a bill for.
        :param bill_range:          Bill range
        '''

        logging.debug('Lade %s ...', path)
        workbook = openpyxl.load_workbook(filename=str(path))

        self.invoices = []
        for row in self.__get_rows(workbook, InputSheet.INVOICES):
            try:
                invoice_range = DateRange(Date.from_str(row[5]), Date.from_str(row[6]))
                if invoice_range.overlaps(bill_range):
                    invoice = Invoice(
                        row[0],
                        row[1],
                        row[2],
                        Date.from_str(row[3]),
                        row[4],
                        invoice_range,
                        row[7],
                        row[8],
                        row[9],
                        row[11])
                    self.invoices.append(invoice)
            except ValueError as value_error:
                raise ValueError(f'Lesen von {row} fehlgeschlagen: {str(value_error)}') 
        logging.debug('%d Rechnungen', len(self.invoices))

        self.appartements = []
        for row in self.__get_rows(workbook, InputSheet.APPARTEMENTS):
            self.appartements.append(Appartement(row[0], row[1]))
        logging.debug('%d Wohnungen', len(self.appartements))

        self.appartement = next((a for a in self.appartements if a.name == appartement_name), None)
        if not self.appartement:
            logging.error('Gesuchte Wohnung nicht in Eingabedatei!')
            raise InputFileError()

        self.tenants = []
        for row in self.__get_rows(workbook, InputSheet.TENANTS):
            try:
                # For 'moving out' we also accept None
                moving_out = None
                if row[3]:
                    moving_out = Date.from_str(row[3])

                self.tenants.append(
                    Tenant(
                        row[0],
                        row[1],
                        Date.from_str(row[2]),
                        moving_out,
                        int(row[4]),
                        row[5],
                        row[6]
                    )
                )
            except ValueError as value_error:
                raise ValueError(f'Lesen von {row} fehlgeschlagen: {str(value_error)}') 
        logging.debug('%d Mieter', len(self.tenants))

        self.tenant = next((t for t in self.tenants if bill_range.begin in t\
            and t.appartement == appartement_name), None)
        if not self.tenant:
            logging.error('Kein Mieter für den Zeitraum in der Wohnung bekannt.')
            raise InputFileError()
        if bill_range.end not in self.tenant:
            logging.error('Der Mieter war nicht über den gesamten Rechnungszeitraum'\
                ' in der Wohnung (%s bis %s)', self.tenant.moving_in, self.tenant.moving_out)
            raise InputFileError()
        logging.info('Mieter: %s', self.tenant.name)

        self.meter_values = []
        for row in self.__get_rows(workbook, InputSheet.METER_VALUES):
            try:
                self.meter_values.append(MeterValue(row[0], row[1], Date.from_str(row[2]), row[3]))
            except ValueError as value_error:
                raise ValueError(f'Lesen von {row} fehlgeschlagen: {str(value_error)}') 
        logging.debug('%d Zählerstände', len(self.meter_values))

        self.meter = []
        for row in self.__get_rows(workbook, InputSheet.METERS):
            self.meter.append(Meter(row[0], row[1], row[2]))
        logging.debug('%d Zähler', len(self.meter))

        self.bcis = []
        for row in self.__get_rows(workbook, InputSheet.BILL_CALCULATION_ITEMS):
            bci = BillCalculationItem(row[0], row[1], row[2], row[3], row[4], row[5])
            # Skip BCIs that are not relevant for this appartement
            if bci.appartement == appartement_name:
                # Check that for consumption there is a unit specified
                if bci.split == SplitType.PER_CONSUMPTION.value and not bci.unit:
                    logging.critical('Abrechnung für "%s", "%s" nach Verbrauch, aber ohne Einheit zu definieren',
                        bci.appartement, bci.invoice_type)
                    raise InputFileError
                self.bcis.append(bci)
        logging.debug('%d Rechnungseinstellungen', len(self.bcis))

    def __get_rows(self, workbook, sheet: InputSheet):
        ''' Get all data rows of given sheet. '''
        sheet = workbook[sheet.value]
        # We need to filter out the rows without content, because we will receive those as well.
        return filter(lambda row: row[0], sheet.iter_rows(min_row=2, values_only=True))

    def get_meter(self, meter_name) -> Meter:
        ''' Get Meter by name. '''
        return next((m for m in self.meter if m.name == meter_name), None)

    def get_invoices(self, bci: BillCalculationItem, date_range: DateRange) -> List[Invoice]:
        ''' List all invoices related to given bci in given date range. '''
        return filter(lambda i: i.type == bci.invoice_type and i.range.overlaps(date_range),
            self.invoices)

class ResultSheet(enum.Enum):
    ''' Sheet names in the result sheet. '''
    OVERVIEW = 'Zusammenfassung'
    DETAILS = 'Details'
    METER_VALUES = 'Zählerstände'

class CellWriter:
    ''' Helper class to write into a cell of a sheet '''

    def __init__(self, sheet, row: int, column: int):
        self._sheet = sheet
        self._row = row
        self._column = column

    def write_date(self, date: Date, style: str = 'content'):
        ''' Write a date '''
        self.write(date.date, style=style, number_format='DD.MM.YY')

    def write_number(self, number: str, style: str = 'content', unit: str = None, precision: int = 2):
        ''' Write a nunmber '''
        precision_format = '0'
        if precision > 0:
            precision_format += '.' + (precision * '0')

        if unit:
            self.write(number, style=style, number_format=f'{precision_format}" {unit}"')
        else:
            self.write(number, style=style, number_format=precision_format)

    def write_currency(self, number: str, style: str = 'content'):
        ''' Write a currency '''
        self.write(number, style=style, number_format='0.00" "€')

    def write_percentage(self, number: str, style: str = 'content'):
        ''' Write a percentage '''
        self.write(number, style=style, number_format='0.00" "%')

    def write(self, content, style: str = 'content', number_format = None):
        ''' Write into a cell '''
        cell = self._sheet.cell(row=self._row, column=self._column)
        cell.style = style
        cell.value = content
        if number_format:
            cell.number_format = number_format

    def row(self):
        ''' Get the row of this writer '''
        return self._row

class RowWriter(CellWriter):
    ''' Helper class to write a row into a sheet '''

    def write(self, content, style: str = 'content', number_format = None):
        ''' Write into a cell '''
        super().write(content, style=style, number_format=number_format)
        self._column += 1

class ResultSheetWriter:
    '''
    Write the result Excel sheet.
    '''

    template_filepath = 'example-bill.xlsx'

    def __init__(self):
        self._current_row = {
            ResultSheet.METER_VALUES: 2,
            ResultSheet.DETAILS: 2,
            ResultSheet.OVERVIEW: 5  # Write categories
        }

        self._wb = openpyxl.load_workbook(self.template_filepath)
        self.__define_styles()

    def __define_styles(self):
        ''' Create named styles for use in CellWrite, RowWriter '''
        style = openpyxl.styles.NamedStyle(name='header')
        style.font = openpyxl.styles.Font(name='Calibri', bold=True, size=11)
        self._wb.add_named_style(style)

        style = openpyxl.styles.NamedStyle(name='header-overview')
        style.font = openpyxl.styles.Font(name='Calibri', bold=True, size=14)
        self._wb.add_named_style(style)

        style = openpyxl.styles.NamedStyle(name='double-underlined')
        style.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True, underline='double')
        self._wb.add_named_style(style)

        style = openpyxl.styles.NamedStyle(name='content')
        style.font = openpyxl.styles.Font(name='Calibri', size=11)
        self._wb.add_named_style(style)

        style = openpyxl.styles.NamedStyle(name='bold')
        style.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
        self._wb.add_named_style(style)

    def row_writer(self, sheet: ResultSheet):
        ''' Create a row writer for the next row in given sheet. '''
        writer = RowWriter(self._wb[sheet.value], self._current_row[sheet], 1)
        self._current_row[sheet] += 1
        return writer

    def cell_writer(self, sheet: ResultSheet, row: int, column: int):
        ''' Create a cell writer for the given row and column in given sheet. '''
        return CellWriter(self._wb[sheet.value], row, column)

    def write_overview(self, appartement_name, tenant_name, bill_range: DateRange, bcis):
        ''' Write information in overview sheet '''

        # Title
        self.cell_writer(ResultSheet.OVERVIEW, 1, 2).write('Nebenkostenabrechnung ' + appartement_name, style='header-overview')
        self.cell_writer(ResultSheet.OVERVIEW, 1, 4).write(tenant_name, style='header-overview')

        # Date range
        self.cell_writer(ResultSheet.OVERVIEW, 2, 3).write_date(bill_range.begin)
        self.cell_writer(ResultSheet.OVERVIEW, 2, 4).write_date(bill_range.end)

        # Invoice type sums
        invoice_types = sorted(set([bci.invoice_type for bci in bcis]))
        data_end_row = 0
        for invoice_type in invoice_types:
            row = self.row_writer(ResultSheet.OVERVIEW)
            row.write('')
            row.write(invoice_type)

            row.write_currency(f'=D{row.row()}/((_xlfn.days($D$2,$C$2)+1)/IF(OR(MOD($C$2,400)=0,AND(MOD($C$2,4)=0,MOD($C$2,100)<>0)),365,366)*12)')
            row.write_currency(f'=SUMIF({ResultSheet.DETAILS.value}!$D$2:$D${self._current_row[ResultSheet.DETAILS]},"{invoice_type}",{ResultSheet.DETAILS.value}!$N$2:$N${self._current_row[ResultSheet.DETAILS]})')

            data_end_row = row.row()

        row = self.row_writer(ResultSheet.OVERVIEW)  # empty row

        # SUMS
        row = self.row_writer(ResultSheet.OVERVIEW)
        row.write('')
        row.write('Summe', style='bold')
        row.write_currency(f'=SUM($C$5:$C${row.row()-2}', style='bold')
        row.write_currency(f'=SUM($D$5:$D${row.row()-2}', style='bold')
        row_sums = row.row()

        # Payments on advance
        row = self.row_writer(ResultSheet.OVERVIEW)
        row.write('')
        row.write('Abschlagszahlungen', style='bold')
        row.write('')
        row.write_currency('BITTE EINTRAGEN', style='bold')
        row_payments = row.row()

        row = self.row_writer(ResultSheet.OVERVIEW)  # empty row

        # Result
        row = self.row_writer(ResultSheet.OVERVIEW)
        row.write('')
        row.write('Ergebnis', style='bold')
        row.write('')
        row.write_currency(f'=$D${row_payments}-$D${row_sums}', style='double-underlined')

        # Pie Chart
        sheet = self._wb[ResultSheet.OVERVIEW.value]
        pie = openpyxl.chart.PieChart()
        labels = openpyxl.chart.Reference(sheet, min_col=2, min_row=5, max_row=data_end_row)
        data = openpyxl.chart.Reference(sheet, min_col=4, min_row=5, max_row=data_end_row)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.height = 19 # default is 7.5
        pie.width = 21 # default is 15
        
        pie.legend = None

        pie.series[0].graphicalProperties.line.solidFill = "FFFFFF"
        pie.series[0].graphicalProperties.line.width = 20000

        # Show values in chart
        pie.dataLabels = openpyxl.chart.label.DataLabelList()
        pie.dataLabels.showSerName = False   # Show percentage
        pie.dataLabels.showVal = False
        pie.dataLabels.dLblPos = "outEnd"
        
        sheet.add_chart(pie, f'B{row.row() + 2}')

    def save(self, filepath):
        ''' Write the result '''

        logging.info('Schreibe Rechnung nach %s ...', filepath)
        self._wb.save(filepath)
